import pdfplumber
import pandas as pd
import re
from pathlib import Path
from openpyxl.styles import PatternFill
from datetime import datetime

# ---------- 辅助函数 ----------
def get_company_from_filename(pdf_path):
    name = pdf_path.stem
    if '.' in name and name.split('.')[0].isdigit():
        return name.split('.', 1)[1]
    return name

def parse_date(date_str):
    """将各种日期格式转为datetime对象，支持 '2026年05月22日'、'2026-05-22'、'2026/05/22'"""
    if not date_str:
        return None

    for old, new in [('年', '-'), ('月', '-'), ('日', ''), ('/', '-')]:
        date_str = date_str.replace(old, new)

    date_str = date_str.strip()
    try:
        return datetime.strptime(date_str, '%Y-%m-%d')
    except:
        try:
            return datetime.strptime(date_str, '%Y-%m')
        except:
            return None

def extract_invoice_data(pdf_path):
    data = {
        '发票号码': '',
        '开票日期': '',
        '开票金额': '',
        '开票公司': ''
    }
    try:
        with pdfplumber.open(pdf_path) as pdf:
            full_text = ''
            for page in pdf.pages:
                text = page.extract_text()
                if text:
                    full_text += text + '\n'

            if not full_text.strip():
                data['开票公司'] = get_company_from_filename(pdf_path) + '(图片需OCR)'
                return data

            # 1. 发票号码
            match = re.search(r'发票号码[：:]\s*(\w+)', full_text)
            if match:
                data['发票号码'] = match.group(1)

            # 2. 开票日期
            match = re.search(r'开票日期[：:]\s*(\d{4}[年/-]\d{1,2}[月/-]\d{1,2}[日]?)', full_text)
            if match:
                data['开票日期'] = match.group(1)

            # 3. 开票金额（多级匹配）
            match = None
            # 优先匹配价税合计
            match = re.search(r'价税合计[（(]小写[）)]?[：:]*\s*[￥¥]?([\d,]+\.?\d*)', full_text, re.DOTALL)
            if not match:
                match = re.search(r'(?:合计|总计)[：:]*\s*[￥¥]?([\d,]+\.?\d*)', full_text, re.DOTALL)
            if not match:
                # 兜底：取所有金额中最大的（过滤小于100的单价）
                all_amounts = re.findall(r'[￥¥]?([\d,]+\.\d{2})', full_text)
                valid = []
                for amt in all_amounts:
                    try:
                        num = float(amt.replace(',', ''))
                        if num > 100:
                            valid.append(num)
                    except:
                        pass
                if valid:
                    data['开票金额'] = str(max(valid))
            if match:
                data['开票金额'] = match.group(1).replace(',', '')

            # 4. 开票公司
            match = re.search(r'销[售]?方名称[：:]\s*(.+)', full_text)
            if match:
                data['开票公司'] = match.group(1).strip()
            else:
                data['开票公司'] = get_company_from_filename(pdf_path)

    except Exception as e:
        print(f'  ⚠️ 出错: {pdf_path.name} - {e}')
        if not data['开票公司']:
            data['开票公司'] = get_company_from_filename(pdf_path)
    return data

# ---------- 主程序 ----------
if __name__ == '__main__':
    work_dir = Path(__file__).parent
    pdf_list = list(work_dir.glob('*.pdf'))

    if not pdf_list:
        print(f'❌ 没找到PDF文件！请把本脚本和PDF放在同一个文件夹：{work_dir}')
        exit()

    print(f'📂 共发现 {len(pdf_list)} 个PDF文件，开始提取...\n')

    results = []
    for i, pdf_file in enumerate(pdf_list, 1):
        print(f'  处理中 ({i}/{len(pdf_list)}): {pdf_file.name}')
        row = extract_invoice_data(pdf_file)
        results.append(row)

    df = pd.DataFrame(results)
    df = df[['发票号码', '开票日期', '开票金额', '开票公司']]

    # ---------- 写入 Excel 并应用颜色 ----------
    output_file = work_dir / '发票台账_带标黄.xlsx'
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Sheet1')
        workbook = writer.book
        worksheet = writer.sheets['Sheet1']

        # 找到列索引（注意：A列是索引？但我们不写索引，所以列名从A开始）
        # 列顺序：A=发票号码，B=开票日期，C=开票金额，D=开票公司
        date_col = None
        amount_col = None
        for col_idx, col_name in enumerate(df.columns, 1):  # 1-based
            if col_name == '开票日期':
                date_col = col_idx
            if col_name == '开票金额':
                amount_col = col_idx

        if date_col and amount_col:
            cutoff = datetime(2026, 3, 16)
            # 遍历数据行（从第2行开始，第1行是标题）
            for row_idx in range(2, len(df) + 2):
                date_cell = worksheet.cell(row=row_idx, column=date_col)
                date_str = date_cell.value
                if date_str:
                    dt = parse_date(date_str)
                    if dt and dt < cutoff:  # 如果日期在 2026-06-16 之前
                        amount_cell = worksheet.cell(row=row_idx, column=amount_col)
                        amount_cell.fill = yellow_fill
        else:
            print('⚠️ 未找到日期或金额列，跳过标黄。')

    print(f'\n✅ 大功告成！共处理 {len(results)} 张发票。')
    print(f'📁 结果文件：{output_file}')
    print('\n📊 完整预览（仅显示前5行）：')
    print(df.head())
    print('\n💡 提示：开票日期在 2026-03-16 之前的，金额单元格已标黄。')